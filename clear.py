import openpyxl

workbook = openpyxl.load_workbook('static/Left Leg.xlsx')  # create file
worksheets = workbook.sheetnames  # create worksheet
for sheet in worksheets:
    worksheet = workbook[sheet]
    worksheet.delete_rows(1, worksheet.max_row)

workbook.save('static/Left Leg.xlsx')

workbook = openpyxl.load_workbook('static/Right Leg.xlsx')  # create file
worksheets = workbook.sheetnames  # create worksheet
for sheet in worksheets:
    worksheet = workbook[sheet]
    worksheet.delete_rows(1, worksheet.max_row)

workbook.save('static/Right Leg.xlsx')

dict = {
        1: 'static/left_angle.txt',
        2: 'static/right_angle.txt',
        3: 'static/left_position_hip.txt',
        4: 'static/left_position_knee.txt',
        5: 'static/left_position_ankle.txt',
        6: 'static/left_position_hip.txt',
        7: 'static/left_position_knee.txt',
        8: 'static/left_position_ankle.txt',
        9: 'static/right_position_hip.txt',
        10: 'static/right_position_knee.txt',
        11: 'static/right_position_ankle.txt',
        12: 'static/right_position_hip.txt',
        13: 'static/right_position_knee.txt',
        14: 'static/right_position_ankle.txt',
        15: 'static/left_velocity.txt',
        16: 'static/right_velocity.txt',
    }

for value in dict.values():
    open(value, 'w').close()
